import sys 
import csv 
import glob 
import codecs
import os
import datetime as dt
import errno
import re
from collections import defaultdict
import openpyxl


"""
takes 3 arguments
1. jumpserver log folder
2. targeted server log folder
3. potential user list 

ex)
py PredictOper_2.py "path\to\jumpServerLog" "path\to\serferfolder" "userlist.xlsx"

"""

class PredictOper:
    oper = "oper"
    notOper = -1
    operList = []
    emptyList = []
    JUMP_HEADER = ["No", "管理No", "検証結果", "host", "user", "pid", "src_ip", "src_port", "Accepted", "", "sessionOpened", "", "sessionClosed", "", "Company","Oper?"]
    ETL_HEADER = ["host","user","pid","src_ip","src_port", "Accepted","sessionClosed"]
    DATEINFO_HEADER = ["NOT FOUND", "NUMBER OF OPER"]
    NOTFOUNDLIST = ["NOT FOUND"]
    dl = [] #datelist
    MAXLOOP = 5 # max times of recursive loop
    
    def getPath(self):
        targetPath ="."
        if len(sys.argv) == 4: # check if there is an argument for the paths
            self.__pathToJumpServerLog = str(sys.argv[1]) + "/*.xlsx"
            self.__pathToETL01 = str(sys.argv[2]) + "/*.csv"
            self.__oper = str(sys.argv[3])
            self.makeOperDict()
            return True
        else:
            return False

    def makeOperDict(self):
        self.operList = defaultdict(list)
        wb = openpyxl.load_workbook(self.__oper)
        ws = wb['list']
        
        for i in range(ws.max_row):
            if i > 0:
                row = ws[i]
                self.operList[row[1].value] = row[2].value

    def updatedate(self, JMPDate):
        """
        ETL_File : C:\TEMP\Splunk_log\長井さん\fdev\ETL01_log\Audit_Report_--_PUJNASETL01-2018-11-25.csv
        JUM file : C:\TEMP\Splunk_log\長井さん\fdev\jumpServerLog\20181118-1124_nix_Jump_Servers_詳細ログレポート_Login-Logout.xlsx
        this function finds a matching pair of ETL_file andf Jump server file
        """
        last = int(JMPDate[0][-1])
        second = int(JMPDate[0][-2])
        if last != 9:
            last += 1
            JMPDate[0] = JMPDate[0][1:3] + "-" + JMPDate[0][3:4] + str(last)
        else:
            second += 1
            JMPDate[0] = JMPDate[0][1:3] + "-" + str(second) + "0"
        return JMPDate[0]

    def trimText(self, JMPFILE):
        findDate = re.compile('(-\d{4})') #regx find '-(int)(int)(int)(int)'
        JMPDate = findDate.findall(JMPFILE)
        return JMPDate
        
    def getFileName(self, file):
        fileArr = file.split("\\" )
        return  "missing_data_" + str(fileArr[-1])
    
    def adjustDateFormat(self, timeStamp):
        timeStamp = timeStamp.replace(r'/', '-')
        if (timeStamp[0]) == '"':
            timeStamp = timeStamp[1:]
        if timeStamp[12] == ":":
            timeStamp = timeStamp[0:11] + "0" + timeStamp[11:12]
        return str(timeStamp[0:13])

    def searchOper(self, potentialOper):
        return self.operList.get(potentialOper)
    
    def dateDictinaryLookUp(self, date):
        return self.dl.get(date)

    def convertTupleToStringFormatForExcel(self, p):
        userID = str(p[4].value.upper())
        operUser = self.searchOper(userID)
        
        stri = (str(p[0].value) + " " + str(p[1].value) + " " + str(p[2].value) + " " + str(p[3].value) + " " + str(p[4].value) + " " + 
                str(p[5].value) + " " + str(p[6].value) + " " + str(p[7].value) + " " + str(p[8].value) + " " + str(p[9].value) + " " +
                str(p[10].value) + " "+ str(p[11].value))

        if operUser == None:
            stri = stri + " No"
        else:
            stri = stri + " " + operUser

        return stri.split()

    def extractTime(self, timeStamp):
        findDate = re.compile('(\d{1,}-\d{1,})')
        JMPDate = findDate.findall(timeStamp)
        return JMPDate[0]

    def dateFormat(self, date):
        # JUMP 2018-10-14 01:01:02 JST
        # ETL  2018/11/12  9:37:35
        trimDate = re.compile('(\d{1,})') #extract only digit
        date_only_num = trimDate.findall(date)
        return date_only_num

    def getDateInfo(self, NOTFOUND, numOfOper):
        l = []
        l.append(NOTFOUND)
        l.append(numOfOper)
        return l

    def openFile(self):
        print("\npath to JumpServer .xlsx file: " + self.__pathToJumpServerLog+ "\n")
        print("\npath to ETL01 .xlsx file: " + self.__pathToETL01+ "\n")
        currentTime = dt.datetime.now().strftime("%Y%m%d")
        fileNum = 1
        for ETL_file in glob.glob(self.__pathToETL01):
            print("\n\n----------- file"+ str(fileNum) + "-----------\n")
            print(ETL_file)

            fileName = "C:\TEMP\Splunk_log\\fdev\jumpServerLog\log_analyse_data\\" + self.getFileName(ETL_file)
            ETL_date = fileName[-9:-4]
            
            for JMPFILE in glob.glob(self.__pathToJumpServerLog):
                JMPDate = self.trimText(JMPFILE)
                if len(JMPDate) == 1:
                    JMPDate = self.updatedate(JMPDate)

                if JMPDate == ETL_date: # find the match based on the date
                    print(str(ETL_date) + " "  + str(JMPDate))
                    print(ETL_file)
                    print(JMPFILE)
                    self.dl = self.makeDateList(JMPFILE)
                    self.findPotentialOperUser(ETL_file, JMPFILE, fileNum)
            fileNum+=1

    def makeDateList(self, JMPFILE):
        """
         builds a dictionary based on the "YYYY-MM-DD HH"
        """
        d = defaultdict(list)
        self.dl = []
        #testcase = "C:\TEMP\Splunk_log\長井さん\\fdev\\testcase\dictlist_testcase.xlsx"
        #testcase = "C:\TEMP\Splunk_log\長井さん\\fdev\jumpServerLog\\20181028-1103_nix_Jump_Servers_詳細ログレポート_Login-Logout.xlsx"

        # open Jump server .xlsx
        # loop thru JMP server each row
        
        #wb = openpyxl.load_workbook(testcase)
        #ws = wb['Audit_Report_--_nix_Jump_Server'] #'Audit_Report_--_nix_Jump_Server'
        #ws = wb['Sheet1']

        wb = openpyxl.load_workbook(JMPFILE)
        ws = wb['Audit_Report_--_nix_Jump_Server']

        for i in range(ws.max_row):
            if i % 500 == 0:
                print(str(i) + "/" + str(ws.max_row))
            
            if i != 0: # ignore the firt time
                row = ws[i]
                timeRange = str(str(row[8].value)[0:13]) #yyyy-mm-dd tt
                d[timeRange].append(row)
        print("dictinary done")
        return d
    
    def findPotentialOperUser(self, ETL_file, JMPFILE, fileNum):
            
        """ file to save """
        wb_save = openpyxl.Workbook()
        ws_save = wb_save.active
        dataInfo =[]
        NOTFOUND = 0
        numOfOper = 0

        # open ELT_FILE
        with codecs.open(ETL_file,"r",encoding='utf-8', errors='ignore') as fopen:
            csvread = csv.reader(fopen,delimiter='\t') # grab all the contents from the fil
            for row in csvread:
                r = str(row).split(",")
                potentialOperUserlist = self.findJumpServerLog(r, JMPFILE)

                if potentialOperUserlist != self.notOper:
                    """ write date into an excel file """
                    ws_save.append(self.ETL_HEADER)
                    ws_save.append(r) # oper data
                    ws_save.append(self.JUMP_HEADER)
                    
                    if len(potentialOperUserlist) == 0:
                        NOTFOUND += 1
                        ws_save.append(self.NOTFOUNDLIST)
                    else:
                        numOfOper += 1
                        for potenUser in potentialOperUserlist:
                            potenUser = self.convertTupleToStringFormatForExcel(potenUser)
                            ws_save.append(potenUser)
                    ws_save.append(self.emptyList)
            
            """ Append data info at the end of the excel file"""
            dataInfo = self.getDateInfo(NOTFOUND, numOfOper)
            ws_save.append(self.DATEINFO_HEADER)
            ws_save.append(dataInfo)

        filesave = "C:\TEMP\Splunk_log\長井さん\\fdev\match\\" + "potential_oper_user_" + str(self.extractTime(JMPFILE)) + ".xlsx"
        wb_save.save(filesave)
        print(filesave + "saved!")

    def findJumpServerLog(self, row, JMPFILE): #row is ETLrow

        """ ETL_accepted ex) 2018-10-23 17 """
        ETL_accepted = str(row[-2])
        ETL_sessionClosed = row[-1]
        ETL_user = row[1]
        potentialOperUserlist = self.notOper

        """ if user is not oper """
        if ETL_user != self.oper:
            return potentialOperUserlist

        """ if user is oper, then continue """
        Ea = self.dateFormat(ETL_accepted)
        Esc = self.dateFormat(ETL_sessionClosed)
        ETL_accepted = self.adjustDateFormat(ETL_accepted)

        if self.dateDictinaryLookUp(ETL_accepted) != None:
            print("------------------------------------")
            potentialOperUserlist = self.findJumpServerLogForOper(row, ETL_accepted, 0)

        return potentialOperUserlist

    def findJumpServerLogForOper(self, ETLrow, ETL_accepted, recurTimes):
        """
            find potential user based on the timeStamp
            if it doesn't find the match, it will recursivly go back and search it up to 4 hours back
        """
        dt_ETL_accepted = 0
        dt_ETL_sessionClosed = 0
        Ea = self.dateFormat(ETLrow[-2])
        Esc = self.dateFormat(ETLrow[-1])
        potentialUser = []
        searchList = self.dateDictinaryLookUp(ETL_accepted) # get the list of the potential users


        print("recurTimes : " +str(recurTimes))
        print(ETL_accepted)

        """ recursion base case """
        if searchList == None or recurTimes == self.MAXLOOP: # go back up to 5 hours prior
            return potentialUser
        
        for i in range(len(searchList)):
            Ja = self.dateFormat(str(searchList[i][-4].value))
            Jsc = self.dateFormat(str(searchList[i][-2].value))

            try:
                if len(Ea) == 6: # YYYY:MM:DD:TT:MM:SS
                    dt_ETL_accepted = dt.datetime(int(Ea[0]), int(Ea[1]), int(Ea[2]), int(Ea[3]), int(Ea[4]), int(Ea[5]))
                    dt_ETL_sessionClosed = dt.datetime(int(Esc[0]), int(Esc[1]), int(Esc[2]), int(Esc[3]), int(Esc[4]), int(Esc[5]))
                else:# YYYY:MM:DD:TT:MM
                    dt_ETL_accepted = dt.datetime(int(Ea[0]), int(Ea[1]), int(Ea[2]), int(Ea[3]), int(Ea[4]))
                    dt_ETL_sessionClosed = dt.datetime(int(Esc[0]), int(Esc[1]), int(Esc[2]), int(Esc[3]), int(Esc[4]))
                
                dt_JMP_accepted = dt.datetime(int(Ja[0]), int(Ja[1]), int(Ja[2]), int(Ja[3]), int(Ja[4]), int(Ja[5]))
                dt_JMP_sessionClosed = dt.datetime(int(Jsc[0]), int(Jsc[1]), int(Jsc[2]), int(Jsc[3]), int(Jsc[4]), int(Ja[5]))
                
                if (dt_ETL_accepted - dt_JMP_accepted ).total_seconds() >= 0 and (dt_JMP_sessionClosed - dt_ETL_sessionClosed).total_seconds() >= 0:
                    potentialUser.append(searchList[i])
            except:
                pass
        
        if len(potentialUser) == 0 and recurTimes <= self.MAXLOOP:
            Upadted_ETL_accepted = self.decrement_ETL_accepted(ETL_accepted)
            potentialUser = self.findJumpServerLogForOper(ETLrow, Upadted_ETL_accepted, (recurTimes+1))

        print(len(potentialUser))
        return potentialUser
    
    def decrement_ETL_accepted(self, ETL_accepted):
        """ 
            ETL_accepted = 2018-10-23 05
            time extract " 05" 
            date extract "2018-10-23"
        """
        time = re.compile('( \d{1,2})')
        date = re.compile('(\d{4}-\d{1,2}-\d{1,2})')
        
        _time = int(time.findall(ETL_accepted)[0].replace(" ", ""))
        _time -= 1
        _date = date.findall(ETL_accepted)[0] 
        decremented_ETL_accepted = _date + " " + str(_time)

        return decremented_ETL_accepted

    def Error_noPath(self):
        print("*******************************************************************************\n") 
        print("*** Please enter a path to the .csv file you would like to covert into .exl ***\n")
        print("*******************************************************************************\n")


""" main """
if __name__ == "__main__":

    p = PredictOper()

    if(p.getPath()):
        p.openFile()
    else:
        p.Error_noPath()
